from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from uuid import NAMESPACE_URL, uuid5

import openpyxl
import pytest

from src.auth import User
from src.models.enums import Role
from src.payments.finance import (
    BankTransaction,
    PaymentBatchService,
    PaymentBatchStatus,
    PaymentExportService,
    PaymentReadiness,
    PaymentReadinessService,
    PaymentReconciliationService,
    PaymentRecord,
    PaymentRepository,
    PaymentStateMachine,
    PaymentStatus,
    ReconciliationStatus,
    classify_rib,
    mask_rib,
    payment_reference,
    payment_report,
    rib_fingerprint,
)
from src.restaurants.registry_models import RibValueStatus

ADMIN = User("admin", "Admin", "admin@example.test", Role.ADMIN)
FINANCE = User("finance", "Finance", "finance@example.test", Role.FINANCE)
VIEWER = User("viewer", "Viewer", "viewer@example.test", Role.VIEWER)
VALID_RIB = "001122334455667788990011"


def source_record(rid: str = "R1", rib: str | None = VALID_RIB, amount: str = "100.00"):
    restaurant = SimpleNamespace(
        restaurant_name=f"Restaurant {rid}",
        rib=rib,
        bank="Test Bank",
        legal_master_review_status=None,
        readiness=SimpleNamespace(identity_ready=True),
    )
    settlement = SimpleNamespace(
        restaurant_id=rid,
        period_code="2026-07-P2",
        financial_policy_version="cashco_legacy_v1",
        sales_ttc=Decimal("120.00"),
        invoice_ttc=Decimal("20.00"),
        net_payable=Decimal(amount),
        settlement_status=SimpleNamespace(value="READY"),
    )
    return restaurant, settlement


def record(rid: str = "R1", amount: str = "100.00") -> PaymentRecord:
    restaurant, settlement = source_record(rid, amount=amount)
    return PaymentReadinessService().evaluate(
        restaurant, settlement, documents_ready=True
    )


def test_payment_readiness_rib_validation_mask_and_fingerprint() -> None:
    service = PaymentReadinessService()
    valid = service.evaluate(*source_record(), documents_ready=True)
    missing = service.evaluate(*source_record("R2", None), documents_ready=True)
    invalid = service.evaluate(*source_record("R3", "123"), documents_ready=True)
    assert valid.payment_readiness == PaymentReadiness.PAYMENT_READY
    assert missing.payment_readiness == PaymentReadiness.RIB_MISSING
    assert invalid.payment_readiness == PaymentReadiness.RIB_INVALID
    assert classify_rib(VALID_RIB) == RibValueStatus.VALID_FORMAT
    assert mask_rib(VALID_RIB).endswith("0011") and VALID_RIB not in mask_rib(VALID_RIB)
    assert rib_fingerprint(VALID_RIB) == rib_fingerprint(
        "0011 2233 4455 6677 8899 0011"
    )
    assert VALID_RIB not in valid.model_dump_json()


def test_payment_specific_blockers_are_independent_from_documents() -> None:
    restaurant, settlement = source_record()
    no_document = PaymentReadinessService().evaluate(
        restaurant, settlement, documents_ready=False
    )
    conflict = PaymentReadinessService().evaluate(
        SimpleNamespace(
            **{**restaurant.__dict__, "legal_master_review_status": "CONFLICT"}
        ),
        settlement,
        documents_ready=True,
    )
    review = PaymentReadinessService().evaluate(
        restaurant,
        SimpleNamespace(
            **{
                **settlement.__dict__,
                "settlement_status": SimpleNamespace(value="REVIEW_REQUIRED"),
            }
        ),
        documents_ready=True,
    )
    assert no_document.payment_readiness == PaymentReadiness.DOCUMENT_NOT_READY
    assert conflict.payment_readiness == PaymentReadiness.LEGAL_SOURCE_CONFLICT
    assert review.payment_readiness == PaymentReadiness.SETTLEMENT_REVIEW_REQUIRED


def test_payment_state_machine_and_paid_evidence_rbac() -> None:
    machine = PaymentStateMachine()
    item = record()
    with pytest.raises(PermissionError):
        machine.transition(item, PaymentStatus.IN_BATCH, VIEWER)
    in_batch = machine.transition(item, PaymentStatus.IN_BATCH, FINANCE)
    validated = machine.transition(in_batch, PaymentStatus.FINANCE_VALIDATED, FINANCE)
    pending = machine.transition(validated, PaymentStatus.PAYMENT_PENDING, FINANCE)
    with pytest.raises(ValueError):
        machine.transition(pending, PaymentStatus.PAID, FINANCE)
    paid = machine.transition(
        pending,
        PaymentStatus.PAID,
        FINANCE,
        bank_reference="BANK-1",
        executed_payment_date=date(2026, 8, 20),
    )
    assert paid.bank_reference == "BANK-1"
    with pytest.raises(ValueError):
        machine.transition(item, PaymentStatus.PAID, ADMIN)


def test_batch_snapshot_totals_stale_validation_and_determinism() -> None:
    items = (record("R2", "20.25"), record("R1", "10.25"))
    service = PaymentBatchService()
    batch = service.preview(items, FINANCE)
    assert batch.total_net_payable == Decimal("30.50")
    assert batch.batch_id == service.preview(reversed(items), FINANCE).batch_id
    with pytest.raises(ValueError):
        service.validate(batch, FINANCE, "wrong")
    validated = service.validate(
        batch, FINANCE, f"VALIDATE PAYMENT {batch.batch_id}", items
    )
    assert validated.status == PaymentBatchStatus.VALIDATED
    changed = items[0].model_copy(update={"settlement_snapshot_hash": "changed"})
    stale = service.validate(
        batch, ADMIN, f"VALIDATE PAYMENT {batch.batch_id}", (changed, items[1])
    )
    assert stale.status == PaymentBatchStatus.PAYMENT_BATCH_STALE
    with pytest.raises(PermissionError):
        service.preview(items, VIEWER)


def test_finance_export_is_user_triggerable_in_memory_and_contains_no_raw_rib() -> None:
    batch = PaymentBatchService().preview((record(),), FINANCE)
    payload = PaymentExportService().workbook(batch)
    workbook = openpyxl.load_workbook(BytesIO(payload))
    assert workbook.sheetnames == [
        "PAYMENTS",
        "SUMMARY",
        "BLOCKED",
        "RECONCILIATION",
        "README",
    ]
    assert VALID_RIB.encode() not in payload
    assert workbook["SUMMARY"]["D2"].value == 100
    with pytest.raises(PermissionError):
        PaymentExportService().workbook(batch, user=VIEWER, full_ribs={"R1": VALID_RIB})
    controlled = PaymentExportService().workbook(
        batch, user=FINANCE, full_ribs={"R1": VALID_RIB}
    )
    assert (
        openpyxl.load_workbook(BytesIO(controlled))["PAYMENTS"]["E2"].value == VALID_RIB
    )


def test_reconciliation_exact_partial_overpayment_unmatched_and_duplicate() -> None:
    item = record()
    service = PaymentReconciliationService()
    transactions = (
        BankTransaction(
            transaction_date=date(2026, 8, 20),
            amount=Decimal("100.00"),
            bank_reference="B1",
            internal_reference=item.payment_reference,
        ),
        BankTransaction(
            transaction_date=date(2026, 8, 20),
            amount=Decimal("40.00"),
            bank_reference="B2",
            rib=VALID_RIB,
        ),
        BankTransaction(
            transaction_date=date(2026, 8, 20),
            amount=Decimal("120.00"),
            bank_reference="B3",
            internal_reference=item.payment_reference,
        ),
        BankTransaction(
            transaction_date=date(2026, 8, 20),
            amount=Decimal("1.00"),
            bank_reference="B4",
        ),
        BankTransaction(
            transaction_date=date(2026, 8, 20),
            amount=Decimal("1.00"),
            bank_reference="B4",
        ),
    )
    results = service.reconcile((item,), transactions)
    assert results[0].status == ReconciliationStatus.EXACT_MATCH
    assert (
        results[1].status == ReconciliationStatus.MANUAL_REVIEW
    )  # amount mismatch cannot auto-match by RIB
    assert (
        results[2].status == ReconciliationStatus.MANUAL_REVIEW
    )  # same item cannot be matched twice
    assert results[3].status == ReconciliationStatus.MANUAL_REVIEW
    assert results[4].status == ReconciliationStatus.DUPLICATE_TRANSACTION

    partial = service.reconcile(
        (item,),
        (
            BankTransaction(
                transaction_date=date(2026, 8, 20),
                amount=Decimal(40),
                bank_reference="P",
                internal_reference=item.payment_reference,
            ),
        ),
    )[0]
    over = service.reconcile(
        (item,),
        (
            BankTransaction(
                transaction_date=date(2026, 8, 20),
                amount=Decimal(140),
                bank_reference="O",
                internal_reference=item.payment_reference,
            ),
        ),
    )[0]
    assert partial.status == ReconciliationStatus.PARTIAL_PAYMENT
    assert partial.remaining == Decimal("60.00")
    assert over.status == ReconciliationStatus.OVERPAYMENT_REVIEW


def test_references_are_deterministic_and_do_not_claim_bank_identity() -> None:
    assert payment_reference("2026-07-P2", "R1") == payment_reference(
        "2026-07-P2", "R1"
    )
    assert payment_reference("2026-07-P2", "R1").startswith("PAY-2026-07-P2-")
    assert uuid5(NAMESPACE_URL, "x") == uuid5(NAMESPACE_URL, "x")


def test_payment_history_is_append_only_and_audit_contains_no_rib(tmp_path) -> None:
    repository = PaymentRepository(tmp_path / "payments.sqlite3")
    item = record()
    repository.append(item, "PAYMENT_READINESS_RECALCULATED", FINANCE)
    in_batch = PaymentStateMachine().transition(item, PaymentStatus.IN_BATCH, FINANCE)
    repository.append(in_batch, "PAYMENT_BATCH_CREATED", FINANCE)
    assert [entry.version for entry in repository.history(item.payment_id)] == [1, 2]
    audit = repository.audit()
    assert [entry["event"] for entry in audit] == [
        "PAYMENT_READINESS_RECALCULATED",
        "PAYMENT_BATCH_CREATED",
    ]
    assert VALID_RIB not in str(audit)


def test_payment_reporting_uses_payment_state_without_recalculating_amounts() -> None:
    ready = record("R1", "100")
    paid = record("R2", "50").model_copy(
        update={"payment_status": PaymentStatus.PAID}
    )
    reconciled = record("R3", "25").model_copy(
        update={"payment_status": PaymentStatus.RECONCILED}
    )
    report = payment_report((ready, paid, reconciled))
    assert report.net_payable == Decimal(175)
    assert report.paid_amount == Decimal(75)
    assert report.reconciled_amount == Decimal(25)
    assert report.outstanding_amount == Decimal(100)
