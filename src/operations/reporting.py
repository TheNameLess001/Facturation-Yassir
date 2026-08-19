from __future__ import annotations

import csv
from collections.abc import Callable
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.documents.publishing import DocumentPublication, DocumentPublicationStatus
from src.operations.review import ReviewItem
from src.settlement.phase5_models import SettlementSummary


@dataclass(frozen=True)
class FinancialReport:
    period_code: str
    sales_ttc: Decimal
    sales_ht: Decimal
    commission_ht: Decimal
    tva: Decimal
    invoice_ttc: Decimal
    net_payable: Decimal
    restaurant_count: int
    order_count: int

    @property
    def average_invoice_ttc(self) -> Decimal:
        return (
            self.invoice_ttc / self.restaurant_count
            if self.restaurant_count
            else Decimal(0)
        )

    @property
    def average_net_payable(self) -> Decimal:
        return (
            self.net_payable / self.restaurant_count
            if self.restaurant_count
            else Decimal(0)
        )


@dataclass(frozen=True)
class PeriodComparison:
    current_period: str
    comparable_period: str
    absolute_delta: dict[str, Decimal]
    percentage_delta: dict[str, Decimal | None]


@dataclass(frozen=True)
class DocumentReport:
    total_pdfs: int
    invoices: int
    notes_de_debours: int
    statements: int
    versions_created: int
    superseded: int
    failures: int
    already_published: int


class BillingReportingService:
    FINANCIAL_FIELDS = (
        "sales_ttc",
        "sales_ht",
        "commission_ht",
        "tva",
        "invoice_ttc",
        "net_payable",
    )

    @staticmethod
    def financial(summary: SettlementSummary) -> FinancialReport:
        ready = tuple(
            item
            for item in summary.restaurants
            if item.financial_policy_version == "cashco_legacy_v1"
            and item.sales_ttc is not None
        )

        def total(field: str) -> Decimal:
            return sum(
                (getattr(item, field) or Decimal(0) for item in ready), Decimal(0)
            )

        return FinancialReport(
            period_code=summary.period.period_code,
            sales_ttc=total("sales_ttc"),
            sales_ht=total("sales_ht"),
            commission_ht=total("commission_amount"),
            tva=total("invoice_tva"),
            invoice_ttc=total("invoice_ttc"),
            net_payable=total("net_payable"),
            restaurant_count=len(ready),
            order_count=sum(item.total_orders for item in ready),
        )

    @classmethod
    def compare(
        cls, current: FinancialReport, comparable: FinancialReport
    ) -> PeriodComparison:
        absolute: dict[str, Decimal] = {}
        percentage: dict[str, Decimal | None] = {}
        for field in cls.FINANCIAL_FIELDS:
            current_value = getattr(current, field)
            comparable_value = getattr(comparable, field)
            delta = current_value - comparable_value
            absolute[field] = delta
            percentage[field] = (
                delta / comparable_value * Decimal(100)
                if comparable_value != 0
                else None
            )
        return PeriodComparison(
            current_period=current.period_code,
            comparable_period=comparable.period_code,
            absolute_delta=absolute,
            percentage_delta=percentage,
        )

    @staticmethod
    def documents(publications: tuple[DocumentPublication, ...]) -> DocumentReport:
        current = tuple(
            item
            for item in publications
            if item.status
            in {
                DocumentPublicationStatus.PUBLISHED,
                DocumentPublicationStatus.ALREADY_PUBLISHED,
            }
        )
        return DocumentReport(
            total_pdfs=len(current),
            invoices=sum(item.document_type == "INVOICE" for item in current),
            notes_de_debours=sum(
                item.document_type == "NOTE_DE_DEBOURS" for item in current
            ),
            statements=sum(
                item.document_type == "PARTNER_STATEMENT" for item in current
            ),
            versions_created=sum(item.document_version > 1 for item in publications),
            superseded=sum(
                item.status == DocumentPublicationStatus.SUPERSEDED
                for item in publications
            ),
            failures=sum(
                item.status
                in {
                    DocumentPublicationStatus.FAILED,
                    DocumentPublicationStatus.STORAGE_VERIFICATION_FAILED,
                }
                for item in publications
            ),
            already_published=sum(
                item.status == DocumentPublicationStatus.ALREADY_PUBLISHED
                for item in publications
            ),
        )


class BillingExportService:
    def __init__(
        self,
        audit: Callable[[str, dict[str, object]], None] | None = None,
    ) -> None:
        self.audit = audit or (lambda _event, _details: None)

    def period_csv(self, report: FinancialReport) -> bytes:
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(("Metric", "Value"))
        for field in BillingReportingService.FINANCIAL_FIELDS:
            writer.writerow((field, f"{getattr(report, field):.2f}"))
        writer.writerow(("restaurant_count", report.restaurant_count))
        writer.writerow(("order_count", report.order_count))
        self._audit(report.period_code, "Period Billing Summary", "CSV")
        return output.getvalue().encode("utf-8-sig")

    def workbook(
        self,
        report: FinancialReport,
        restaurant_rows: tuple[dict[str, object], ...],
        review_items: tuple[ReviewItem, ...],
        publications: tuple[DocumentPublication, ...],
    ) -> bytes:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Billing Summary"
        summary.append(("Metric", "Value"))
        for field in BillingReportingService.FINANCIAL_FIELDS:
            summary.append(
                (field.replace("_", " ").title(), float(getattr(report, field)))
            )
        summary.append(("Restaurant Count", report.restaurant_count))
        summary.append(("Order Count", report.order_count))
        self._sheet(workbook, "Restaurant Detail", restaurant_rows)
        self._sheet(
            workbook,
            "Review Queue",
            tuple(
                {
                    "Period": item.period,
                    "Restaurant ID": item.restaurant_id,
                    "Restaurant": item.restaurant_name,
                    "Issue Type": item.issue_type.value,
                    "Severity": item.severity.value,
                    "Status": item.status.value,
                    "Description": item.description,
                }
                for item in review_items
            ),
        )
        self._sheet(
            workbook,
            "Document Registry",
            tuple(
                {
                    "Period": item.period_code,
                    "Restaurant ID": item.restaurant_id,
                    "Document Type": item.document_type,
                    "Version": item.document_version,
                    "Status": item.status.value,
                    "Storage": item.provider,
                    "Published At": item.published_at.isoformat()
                    if item.published_at
                    else None,
                }
                for item in publications
            ),
        )
        output = BytesIO()
        workbook.save(output)
        self._audit(report.period_code, "Billing Operations Workbook", "XLSX")
        return output.getvalue()

    def _audit(self, period_code: str, report: str, export_format: str) -> None:
        self.audit(
            "REPORT_EXPORTED",
            {
                "period_code": period_code,
                "report": report,
                "format": export_format,
            },
        )

    @staticmethod
    def _sheet(
        workbook: Workbook,
        title: str,
        rows: tuple[dict[str, object], ...],
    ) -> None:
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(("Status",))
            sheet.append(("No data",))
            return
        headers = tuple(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append(tuple(row.get(header) for header in headers))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="6747E8")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(
                42,
                max(12, max(len(str(cell.value or "")) for cell in column) + 2),
            )
