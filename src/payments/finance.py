from __future__ import annotations

import hashlib
import json
import sqlite3
from collections.abc import Iterable
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal
from enum import StrEnum
from io import BytesIO
from pathlib import Path
from typing import ClassVar
from uuid import NAMESPACE_URL, UUID, uuid5

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict

from src.auth import User
from src.models.enums import Role
from src.restaurants.registry_models import RegisteredRestaurant, RibValueStatus
from src.settlement.phase5_models import RestaurantSettlementEvaluation

MONEY = Decimal("0.01")


class PaymentReadiness(StrEnum):
    PAYMENT_READY = "PAYMENT_READY"
    RIB_MISSING = "RIB_MISSING"
    RIB_INVALID = "RIB_INVALID"
    LEGAL_SOURCE_CONFLICT = "LEGAL_SOURCE_CONFLICT"
    FINANCIAL_NOT_READY = "FINANCIAL_NOT_READY"
    SETTLEMENT_REVIEW_REQUIRED = "SETTLEMENT_REVIEW_REQUIRED"
    DOCUMENT_NOT_READY = "DOCUMENT_NOT_READY"
    PAYMENT_DATA_REVIEW = "PAYMENT_DATA_REVIEW"


class PaymentStatus(StrEnum):
    NOT_READY = "NOT_READY"
    READY = "READY"
    IN_BATCH = "IN_BATCH"
    FINANCE_VALIDATED = "FINANCE_VALIDATED"
    PAYMENT_PENDING = "PAYMENT_PENDING"
    PAID = "PAID"
    RECONCILIATION_REQUIRED = "RECONCILIATION_REQUIRED"
    RECONCILED = "RECONCILED"
    PAYMENT_FAILED = "PAYMENT_FAILED"
    ON_HOLD = "ON_HOLD"
    CANCELLED = "CANCELLED"


class PaymentBatchStatus(StrEnum):
    DRAFT = "DRAFT"
    READY_FOR_REVIEW = "READY_FOR_REVIEW"
    VALIDATED = "VALIDATED"
    EXPORTED = "EXPORTED"
    PAYMENT_PENDING = "PAYMENT_PENDING"
    PARTIALLY_PAID = "PARTIALLY_PAID"
    PAID = "PAID"
    RECONCILED = "RECONCILED"
    CANCELLED = "CANCELLED"
    PAYMENT_BATCH_STALE = "PAYMENT_BATCH_STALE"


class ReconciliationStatus(StrEnum):
    EXACT_MATCH = "EXACT_MATCH"
    MANUAL_REVIEW = "MANUAL_REVIEW"
    PARTIAL_PAYMENT = "PARTIAL_PAYMENT"
    OVERPAYMENT_REVIEW = "OVERPAYMENT_REVIEW"
    DUPLICATE_TRANSACTION = "DUPLICATE_TRANSACTION"


def normalize_rib(value: str | None) -> str | None:
    if not value:
        return None
    compact = "".join(character for character in str(value) if character not in " -")
    return compact or None


def mask_rib(value: str | None) -> str:
    compact = normalize_rib(value)
    return (
        "MISSING" if not compact else f"{'*' * max(0, len(compact) - 4)}{compact[-4:]}"
    )


def rib_fingerprint(value: str | None) -> str | None:
    compact = normalize_rib(value)
    return hashlib.sha256(compact.encode()).hexdigest() if compact else None


def classify_rib(value: str | None) -> RibValueStatus:
    compact = normalize_rib(value)
    if compact is None:
        return RibValueStatus.MISSING
    return (
        RibValueStatus.VALID_FORMAT
        if compact.isdigit() and len(compact) == 24
        else RibValueStatus.INVALID_FORMAT
    )


def payment_reference(period_code: str, restaurant_id: str) -> str:
    short = hashlib.sha256(restaurant_id.encode()).hexdigest()[:10].upper()
    return f"PAY-{period_code}-{short}"


class PaymentRecord(BaseModel):
    model_config = ConfigDict(frozen=True)

    payment_id: UUID
    period_code: str
    restaurant_id: str
    restaurant_name: str
    settlement_snapshot_hash: str
    financial_policy_version: str
    sales_ttc: Decimal
    invoice_ttc: Decimal
    net_payable: Decimal
    currency: str = "MAD"
    rib_status: RibValueStatus
    masked_rib: str
    rib_fingerprint: str | None = None
    bank: str | None = None
    payment_readiness: PaymentReadiness
    payment_status: PaymentStatus
    payment_batch_id: UUID | None = None
    payment_reference: str
    bank_reference: str | None = None
    planned_payment_date: date | None = None
    executed_payment_date: date | None = None
    reconciled_at: datetime | None = None
    reconciled_by: str | None = None
    created_at: datetime
    updated_at: datetime
    version: int = 1


class PaymentBatch(BaseModel):
    model_config = ConfigDict(frozen=True)

    batch_id: UUID
    period_code: str
    created_at: datetime
    created_by: str
    restaurant_count: int
    total_net_payable: Decimal
    currency: str
    status: PaymentBatchStatus
    snapshot_hash: str
    items: tuple[PaymentRecord, ...]
    validated_at: datetime | None = None
    validated_by: str | None = None
    exported_at: datetime | None = None
    notes: str | None = None


class BankTransaction(BaseModel):
    model_config = ConfigDict(frozen=True)

    transaction_date: date
    amount: Decimal
    bank_reference: str
    beneficiary: str | None = None
    rib: str | None = None
    internal_reference: str | None = None


class ReconciliationResult(BaseModel):
    model_config = ConfigDict(frozen=True)

    bank_reference: str
    payment_id: UUID | None
    status: ReconciliationStatus
    expected: Decimal | None = None
    paid: Decimal
    remaining: Decimal | None = None
    match_method: str | None = None


class PaymentReport(BaseModel):
    model_config = ConfigDict(frozen=True)

    net_payable: Decimal
    payment_ready_amount: Decimal
    paid_amount: Decimal
    reconciled_amount: Decimal
    outstanding_amount: Decimal
    payment_readiness_percent: Decimal
    reconciliation_percent: Decimal


def payment_report(records: Iterable[PaymentRecord]) -> PaymentReport:
    items = tuple(records)
    total = sum((item.net_payable for item in items), Decimal(0))
    ready = sum(
        (
            item.net_payable
            for item in items
            if item.payment_readiness == PaymentReadiness.PAYMENT_READY
        ),
        Decimal(0),
    )
    paid = sum(
        (
            item.net_payable
            for item in items
            if item.payment_status
            in {
                PaymentStatus.PAID,
                PaymentStatus.RECONCILIATION_REQUIRED,
                PaymentStatus.RECONCILED,
            }
        ),
        Decimal(0),
    )
    reconciled = sum(
        (
            item.net_payable
            for item in items
            if item.payment_status == PaymentStatus.RECONCILED
        ),
        Decimal(0),
    )
    return PaymentReport(
        net_payable=total,
        payment_ready_amount=ready,
        paid_amount=paid,
        reconciled_amount=reconciled,
        outstanding_amount=total - paid,
        payment_readiness_percent=(ready / total * 100 if total else Decimal(0)),
        reconciliation_percent=(
            reconciled / total * 100 if total else Decimal(0)
        ),
    )


def _snapshot_hash(settlement: RestaurantSettlementEvaluation) -> str:
    payload = {
        "restaurant_id": settlement.restaurant_id,
        "period": settlement.period_code,
        "policy": settlement.financial_policy_version,
        "sales_ttc": str(settlement.sales_ttc),
        "invoice_ttc": str(settlement.invoice_ttc),
        "net_payable": str(settlement.net_payable),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode()).hexdigest()


class PaymentReadinessService:
    def evaluate(
        self,
        restaurant: RegisteredRestaurant,
        settlement: RestaurantSettlementEvaluation,
        *,
        documents_ready: bool,
        require_documents: bool = True,
    ) -> PaymentRecord:
        rib_status = classify_rib(restaurant.rib)
        if not restaurant.readiness.identity_ready:
            readiness = PaymentReadiness.PAYMENT_DATA_REVIEW
        elif settlement.net_payable is None or settlement.net_payable <= 0:
            readiness = PaymentReadiness.FINANCIAL_NOT_READY
        elif settlement.settlement_status.value != "READY":
            readiness = PaymentReadiness.SETTLEMENT_REVIEW_REQUIRED
        elif require_documents and not documents_ready:
            readiness = PaymentReadiness.DOCUMENT_NOT_READY
        elif restaurant.legal_master_review_status in {"CONFLICT", "BLOCKED"}:
            readiness = PaymentReadiness.LEGAL_SOURCE_CONFLICT
        elif rib_status == RibValueStatus.MISSING:
            readiness = PaymentReadiness.RIB_MISSING
        elif rib_status == RibValueStatus.INVALID_FORMAT:
            readiness = PaymentReadiness.RIB_INVALID
        else:
            readiness = PaymentReadiness.PAYMENT_READY
        now = datetime.now(UTC)
        status = (
            PaymentStatus.READY
            if readiness == PaymentReadiness.PAYMENT_READY
            else PaymentStatus.NOT_READY
        )
        restaurant_id = settlement.restaurant_id
        return PaymentRecord(
            payment_id=uuid5(
                NAMESPACE_URL,
                f"cashco:payment:{settlement.period_code}:{restaurant_id}",
            ),
            period_code=settlement.period_code,
            restaurant_id=restaurant_id,
            restaurant_name=restaurant.restaurant_name or restaurant_id,
            settlement_snapshot_hash=_snapshot_hash(settlement),
            financial_policy_version=settlement.financial_policy_version or "UNKNOWN",
            sales_ttc=settlement.sales_ttc or Decimal(0),
            invoice_ttc=settlement.invoice_ttc or Decimal(0),
            net_payable=settlement.net_payable or Decimal(0),
            rib_status=rib_status,
            masked_rib=mask_rib(restaurant.rib),
            rib_fingerprint=rib_fingerprint(restaurant.rib),
            bank=restaurant.bank,
            payment_readiness=readiness,
            payment_status=status,
            payment_reference=payment_reference(settlement.period_code, restaurant_id),
            created_at=now,
            updated_at=now,
        )


class PaymentStateMachine:
    TRANSITIONS: ClassVar[dict[PaymentStatus, set[PaymentStatus]]] = {
        PaymentStatus.NOT_READY: {PaymentStatus.READY, PaymentStatus.ON_HOLD},
        PaymentStatus.READY: {PaymentStatus.IN_BATCH, PaymentStatus.ON_HOLD},
        PaymentStatus.IN_BATCH: {
            PaymentStatus.FINANCE_VALIDATED,
            PaymentStatus.ON_HOLD,
        },
        PaymentStatus.FINANCE_VALIDATED: {
            PaymentStatus.PAYMENT_PENDING,
            PaymentStatus.ON_HOLD,
        },
        PaymentStatus.PAYMENT_PENDING: {
            PaymentStatus.PAID,
            PaymentStatus.PAYMENT_FAILED,
            PaymentStatus.ON_HOLD,
        },
        PaymentStatus.PAID: {PaymentStatus.RECONCILIATION_REQUIRED},
        PaymentStatus.RECONCILIATION_REQUIRED: {
            PaymentStatus.RECONCILED,
            PaymentStatus.ON_HOLD,
        },
        PaymentStatus.PAYMENT_FAILED: {
            PaymentStatus.ON_HOLD,
            PaymentStatus.PAYMENT_PENDING,
        },
        PaymentStatus.ON_HOLD: {PaymentStatus.READY, PaymentStatus.CANCELLED},
    }

    def transition(
        self,
        record: PaymentRecord,
        target: PaymentStatus,
        user: User,
        **evidence: object,
    ) -> PaymentRecord:
        if user.role not in {Role.ADMIN, Role.FINANCE}:
            raise PermissionError("Finance or Admin role required")
        if target not in self.TRANSITIONS.get(record.payment_status, set()):
            raise ValueError(
                f"Invalid payment transition: {record.payment_status} -> {target}"
            )
        if target == PaymentStatus.PAID and (
            not evidence.get("bank_reference")
            or not evidence.get("executed_payment_date")
        ):
            raise ValueError("Bank reference and execution date are required")
        updates = {
            "payment_status": target,
            "updated_at": datetime.now(UTC),
            "version": record.version + 1,
        }
        if target == PaymentStatus.PAID:
            updates.update(
                bank_reference=str(evidence["bank_reference"]),
                executed_payment_date=evidence["executed_payment_date"],
            )
        return record.model_copy(update=updates)


class PaymentBatchService:
    def preview(
        self, records: Iterable[PaymentRecord], user: User, *, notes: str | None = None
    ) -> PaymentBatch:
        if user.role not in {Role.ADMIN, Role.FINANCE}:
            raise PermissionError("Finance or Admin role required")
        items = tuple(sorted(records, key=lambda item: item.restaurant_id))
        if not items or any(
            item.payment_readiness != PaymentReadiness.PAYMENT_READY for item in items
        ):
            raise ValueError("Every payment item must be payment-ready")
        periods = {item.period_code for item in items}
        if len(periods) != 1:
            raise ValueError("A payment batch must contain exactly one period")
        payload = [
            (
                str(item.payment_id),
                str(item.net_payable),
                item.rib_fingerprint,
                item.settlement_snapshot_hash,
            )
            for item in items
        ]
        digest = hashlib.sha256(
            json.dumps(payload, separators=(",", ":")).encode()
        ).hexdigest()
        period = next(iter(periods))
        batch_id = uuid5(NAMESPACE_URL, f"cashco:payment-batch:{period}:{digest}")
        batch_items = tuple(
            item.model_copy(
                update={
                    "payment_batch_id": batch_id,
                    "payment_status": PaymentStatus.IN_BATCH,
                }
            )
            for item in items
        )
        return PaymentBatch(
            batch_id=batch_id,
            period_code=period,
            created_at=datetime.now(UTC),
            created_by=user.user_id,
            restaurant_count=len(items),
            total_net_payable=sum((item.net_payable for item in items), Decimal(0)),
            currency="MAD",
            status=PaymentBatchStatus.READY_FOR_REVIEW,
            snapshot_hash=digest,
            items=batch_items,
            notes=notes,
        )

    def validate(
        self,
        batch: PaymentBatch,
        user: User,
        confirmation: str,
        current: Iterable[PaymentRecord] | None = None,
    ) -> PaymentBatch:
        if user.role not in {Role.ADMIN, Role.FINANCE}:
            raise PermissionError("Finance or Admin role required")
        if confirmation != f"VALIDATE PAYMENT {batch.batch_id}":
            raise ValueError("Strong payment validation confirmation does not match")
        if current is not None:
            by_id = {item.payment_id: item for item in current}
            if any(
                by_id.get(item.payment_id) is None
                or by_id[item.payment_id].settlement_snapshot_hash
                != item.settlement_snapshot_hash
                or by_id[item.payment_id].rib_fingerprint != item.rib_fingerprint
                for item in batch.items
            ):
                return batch.model_copy(
                    update={"status": PaymentBatchStatus.PAYMENT_BATCH_STALE}
                )
        if batch.total_net_payable != sum(
            (item.net_payable for item in batch.items), Decimal(0)
        ):
            raise ValueError("Payment batch reconciliation failed")
        return batch.model_copy(
            update={
                "status": PaymentBatchStatus.VALIDATED,
                "validated_at": datetime.now(UTC),
                "validated_by": user.user_id,
                "items": tuple(
                    item.model_copy(
                        update={"payment_status": PaymentStatus.FINANCE_VALIDATED}
                    )
                    for item in batch.items
                ),
            }
        )


class PaymentReconciliationService:
    def reconcile(
        self, records: Iterable[PaymentRecord], transactions: Iterable[BankTransaction]
    ) -> tuple[ReconciliationResult, ...]:
        items = tuple(records)
        by_internal = {item.payment_reference: item for item in items}
        by_bank = {item.bank_reference: item for item in items if item.bank_reference}
        seen_references: set[str] = set()
        matched: set[UUID] = set()
        results = []
        for transaction in transactions:
            if transaction.bank_reference in seen_references:
                results.append(
                    ReconciliationResult(
                        bank_reference=transaction.bank_reference,
                        payment_id=None,
                        status=ReconciliationStatus.DUPLICATE_TRANSACTION,
                        paid=transaction.amount,
                    )
                )
                continue
            seen_references.add(transaction.bank_reference)
            item = by_internal.get(transaction.internal_reference or "") or by_bank.get(
                transaction.bank_reference
            )
            method = (
                "INTERNAL_REFERENCE"
                if transaction.internal_reference in by_internal
                else "BANK_REFERENCE"
                if transaction.bank_reference in by_bank
                else None
            )
            if item is None and transaction.rib:
                candidates = [
                    candidate
                    for candidate in items
                    if candidate.rib_fingerprint == rib_fingerprint(transaction.rib)
                    and abs(
                        candidate.net_payable.quantize(MONEY)
                        - transaction.amount.quantize(MONEY)
                    )
                    <= MONEY
                ]
                if len(candidates) == 1:
                    item, method = candidates[0], "RIB_AMOUNT"
            if item is None or item.payment_id in matched:
                results.append(
                    ReconciliationResult(
                        bank_reference=transaction.bank_reference,
                        payment_id=None,
                        status=ReconciliationStatus.MANUAL_REVIEW,
                        paid=transaction.amount,
                    )
                )
                continue
            matched.add(item.payment_id)
            expected = item.net_payable.quantize(MONEY, rounding=ROUND_HALF_UP)
            paid = transaction.amount.quantize(MONEY, rounding=ROUND_HALF_UP)
            status = (
                ReconciliationStatus.EXACT_MATCH
                if paid == expected
                else ReconciliationStatus.PARTIAL_PAYMENT
                if paid < expected
                else ReconciliationStatus.OVERPAYMENT_REVIEW
            )
            results.append(
                ReconciliationResult(
                    bank_reference=transaction.bank_reference,
                    payment_id=item.payment_id,
                    status=status,
                    expected=expected,
                    paid=paid,
                    remaining=max(Decimal(0), expected - paid),
                    match_method=method,
                )
            )
        return tuple(results)


class PaymentExportService:
    def workbook(
        self,
        batch: PaymentBatch,
        blocked: Iterable[PaymentRecord] = (),
        *,
        user: User | None = None,
        full_ribs: dict[str, str] | None = None,
    ) -> bytes:
        if full_ribs and (user is None or user.role not in {Role.ADMIN, Role.FINANCE}):
            raise PermissionError(
                "Finance or Admin role required for a full-RIB export"
            )
        workbook = Workbook()
        payments = workbook.active
        payments.title = "PAYMENTS"
        payments.append(
            (
                "Restaurant ID",
                "Restaurant Name",
                "Period",
                "Net Payable",
                "RIB",
                "Bank",
                "Payment Reference",
                "Batch ID",
                "Status",
            )
        )
        for item in batch.items:
            payments.append(
                (
                    item.restaurant_id,
                    item.restaurant_name,
                    item.period_code,
                    float(item.net_payable),
                    (
                        normalize_rib(full_ribs.get(item.restaurant_id))
                        if full_ribs
                        else "MASKED:" + item.masked_rib
                    ),
                    item.bank,
                    item.payment_reference,
                    str(batch.batch_id),
                    item.payment_status.value,
                )
            )
        summary = workbook.create_sheet("SUMMARY")
        summary.append(
            ("Batch ID", "Period", "Restaurant Count", "Total Net Payable", "Status")
        )
        summary.append(
            (
                str(batch.batch_id),
                batch.period_code,
                batch.restaurant_count,
                float(batch.total_net_payable),
                batch.status.value,
            )
        )
        blocked_sheet = workbook.create_sheet("BLOCKED")
        blocked_sheet.append(
            (
                "Restaurant",
                "Restaurant ID",
                "Net Payable",
                "Blocker",
                "RIB Status",
                "Recommended Action",
            )
        )
        for item in blocked:
            blocked_sheet.append(
                (
                    item.restaurant_name,
                    item.restaurant_id,
                    float(item.net_payable),
                    item.payment_readiness.value,
                    item.rib_status.value,
                    "Resolve payment readiness blocker",
                )
            )
        workbook.create_sheet("RECONCILIATION").append(
            ("Bank Reference", "Expected", "Paid", "Difference", "Status")
        )
        workbook.create_sheet("README").append(
            (
                "SENSITIVE FINANCE EXPORT",
                "User-triggered; never commit or distribute outside Finance.",
            )
        )
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="6417E8")
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class PaymentRepository:
    """Append-only payment history and safe audit registry."""

    def __init__(self, path: Path | str) -> None:
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS payment_history (
                    sequence INTEGER PRIMARY KEY AUTOINCREMENT,
                    payment_id TEXT NOT NULL,
                    payload TEXT NOT NULL,
                    occurred_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS payment_audit (
                    sequence INTEGER PRIMARY KEY AUTOINCREMENT,
                    event TEXT NOT NULL,
                    actor_id TEXT NOT NULL,
                    safe_payload TEXT NOT NULL,
                    occurred_at TEXT NOT NULL
                );
                """
            )

    def append(self, record: PaymentRecord, event: str, actor: User) -> None:
        safe_payload = {
            "payment_id": str(record.payment_id),
            "period_code": record.period_code,
            "restaurant_id": record.restaurant_id,
            "payment_status": record.payment_status.value,
            "payment_readiness": record.payment_readiness.value,
            "version": record.version,
        }
        now = datetime.now(UTC).isoformat()
        with self._connect() as connection:
            connection.execute(
                "INSERT INTO payment_history(payment_id,payload,occurred_at) VALUES(?,?,?)",
                (str(record.payment_id), record.model_dump_json(), now),
            )
            connection.execute(
                "INSERT INTO payment_audit(event,actor_id,safe_payload,occurred_at) VALUES(?,?,?,?)",
                (event, actor.user_id, json.dumps(safe_payload), now),
            )

    def history(self, payment_id: UUID) -> tuple[PaymentRecord, ...]:
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT payload FROM payment_history WHERE payment_id=? ORDER BY sequence",
                (str(payment_id),),
            ).fetchall()
        return tuple(PaymentRecord.model_validate_json(row[0]) for row in rows)

    def audit(self) -> tuple[dict[str, object], ...]:
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT event,actor_id,safe_payload,occurred_at FROM payment_audit ORDER BY sequence"
            ).fetchall()
        return tuple(
            {
                "event": row[0],
                "actor_id": row[1],
                "details": json.loads(row[2]),
                "occurred_at": row[3],
            }
            for row in rows
        )

    def _connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.path)
